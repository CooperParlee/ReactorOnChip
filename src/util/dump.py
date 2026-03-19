import openpyxl
from openpyxl.styles import Font
import datetime as dt

def dump_arrays_to_excel(array1, array2, filename, 
                          headers=("Array 1", "Array 2"), 
                          sheet_name="Data"):
    """Dumps two arrays to an Excel file as columns.

    Args:
        array1: first array of values
        array2: second array of values
        filepath (str): output path e.g. 'output.xlsx'
        headers (tuple): column header names
        sheet_name (str): name of the sheet
    """

    filepath = "logs/excel/" + dt.datetime.now().strftime('%d-%m-%Y_%I-%M_') + filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Headers
    ws["A1"] = headers[0]
    ws["B1"] = headers[1]
    ws["A1"].font = Font(bold=True, name="Arial")
    ws["B1"].font = Font(bold=True, name="Arial")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20

    # Data
    for i, (v1, v2) in enumerate(zip(array1, array2), start=2):
        ws.cell(row=i, column=1, value=v1)
        ws.cell(row=i, column=2, value=v2)

    wb.save(filepath)
    print(f"Saved to {filepath}")


# --- Example usage ---
if __name__ == "__main__":
    time  = [i * 0.005 for i in range(10)]
    Q     = [2228, 2215, 2200, 2190, 2180, 2175, 2160, 2145, 2130, 2110]

    dump_arrays_to_excel(time, Q, "output.xlsx", 
                         headers=("Time (s)", "Heat Flow Q (J)"))